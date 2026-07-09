"""
Pure backend logic for the Riverway Room Schedule Auto-Filler.
No input()/print()-based interaction here on purpose - this module is shared
between the Streamlit app and (optionally) a notebook/script version, and the
UI layer is responsible for collecting answers from the user.
"""
import re
import copy
from datetime import datetime, timedelta

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

FIRST_DATA_COL = 3   # column C
LAST_DATA_COL = 9    # column I
BLOCK_WIDTH = 7       # Mon..Sun
CODE_TO_WEEKDAY = {'M': 0, 'T': 1, 'W': 2, 'R': 3, 'F': 4, 'S': 5}  # anything else -> 6 (Sunday)

GENERIC_STOP = {
    "level", "training", "form", "session", "with", "and", "more", "the", "of", "for", "in", "on", "a", "an",
    "beginner", "master", "practitioner", "teacher", "ft", "day", "evening",
}


def weekday_for_code(ch):
    return CODE_TO_WEEKDAY.get(ch.upper(), 6)


def sig_words(text):
    text = re.sub(r'[^a-zA-Z0-9 ]', ' ', text)
    words = [w.lower() for w in text.split()]
    return [w for w in words if len(w) >= 3 and w not in GENERIC_STOP and not w.isdigit()]


# ----------------------------------------------------------------------------
# ROOM SCHEMA PARSING
# ----------------------------------------------------------------------------
def parse_room_schema(ws):
    max_row = ws.max_row

    def norm(v):
        return "" if v is None else str(v).strip()

    slot_rows = []
    for r in range(1, max_row + 1):
        bval = norm(ws.cell(row=r, column=2).value).lower()
        if bval in ("day", "evening"):
            slot_rows.append((r, bval))

    day_rows = [r for r, t in slot_rows if t == 'day']
    if not day_rows:
        raise ValueError("Could not find any 'Day' rows in column B of the template sheet.")

    blocks = []
    for i, dr in enumerate(day_rows):
        start = dr
        end = (day_rows[i + 1] - 1) if i + 1 < len(day_rows) else max_row
        blocks.append((start, end))

    rooms = []
    for (start, end) in blocks:
        day_row = start
        evening_row = None
        labels = []
        for r in range(start, end + 1):
            bval = norm(ws.cell(row=r, column=2).value).lower()
            if bval == 'evening':
                evening_row = r
            aval = norm(ws.cell(row=r, column=1).value)
            if aval:
                labels.append(aval)
        room_name = labels[0] if labels else f"(unnamed room @ row {start})"
        rooms.append({
            'room_name': room_name,
            'day_row': day_row,
            'evening_row': evening_row,
            'labels': labels,
            'block_start': start,
            'block_end': end,
        })
    return rooms


def suggest_room(course_name, rooms):
    """Pure keyword match - no asking. Returns (room_dict, matched_label) or (None, None)."""
    cwords = set(sig_words(course_name))
    for room in rooms:
        for label in room['labels']:
            lwords = set(sig_words(label))
            if cwords & lwords:
                return room, label
    return None, None


# ----------------------------------------------------------------------------
# DATE / DATA PARSING
# ----------------------------------------------------------------------------
def parse_data_date(value, base_year, fall_start_month=7):
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'year') and hasattr(value, 'month'):
        return value

    s = str(value).strip()
    m = re.match(r'^(\d{1,2})[-\s]([A-Za-z]{3,})$', s)
    if not m:
        raise ValueError(f"Could not parse date value: {value!r}")
    day = int(m.group(1))
    mon_str = m.group(2)[:3].title()
    month = datetime.strptime(mon_str, "%b").month
    year = base_year if month >= fall_start_month else base_year + 1
    return datetime(year, month, day).date()


def load_data_rows(ws_data, base_year):
    headers = {}
    for c in range(1, ws_data.max_column + 1):
        v = ws_data.cell(row=1, column=c).value
        if v:
            headers[str(v).strip().lower()] = c

    required = ['starts', 'ends', 'course', 'name', 'days', 'instructor']
    for req in required:
        if req not in headers:
            raise ValueError(f"Data sheet is missing required column: '{req}'")

    rows = []
    for r in range(2, ws_data.max_row + 1):
        name_val = ws_data.cell(row=r, column=headers['name']).value
        if name_val is None or str(name_val).strip() == "":
            continue

        start_raw = ws_data.cell(row=r, column=headers['starts']).value
        end_raw = ws_data.cell(row=r, column=headers['ends']).value
        start_date = parse_data_date(start_raw, base_year)
        end_date = parse_data_date(end_raw, base_year)

        course_raw = str(ws_data.cell(row=r, column=headers['course']).value or "").strip()
        course_num = re.sub(r'-\d+$', '', course_raw)

        days_raw = str(ws_data.cell(row=r, column=headers['days']).value or "").strip()
        instructor = str(ws_data.cell(row=r, column=headers['instructor']).value or "").strip()
        name = str(name_val).strip()

        notes = ""
        if 'notes' in headers:
            notes = str(ws_data.cell(row=r, column=headers['notes']).value or "").strip()
        reg = ""
        if 'reg' in headers:
            reg = ws_data.cell(row=r, column=headers['reg']).value

        rows.append({
            'row': r,
            'start': start_date,
            'end': end_date,
            'course_num': course_num,
            'name': name,
            'days': days_raw,
            'instructor': instructor,
            'notes': notes,
            'reg': reg,
        })
    return rows


def detect_base_year(template_ws):
    for i in range(1, template_ws.max_row + 1):
        v = template_ws.cell(row=i, column=FIRST_DATA_COL).value
        if hasattr(v, 'year'):
            return v.year
    return datetime.now().year


# ----------------------------------------------------------------------------
# WEEK / COLUMN-BLOCK HELPERS
# ----------------------------------------------------------------------------
def monday_of(d):
    return d - timedelta(days=d.weekday())


def build_week_list(min_start, max_end):
    weeks = []
    cur = monday_of(min_start)
    last_monday = monday_of(max_end)
    while cur <= last_monday:
        weeks.append(cur)
        cur += timedelta(days=7)
    return weeks


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)
        dst_cell.alignment = copy.copy(src_cell.alignment)


def clone_week_block(ws, dst_col_start, src_col_start=FIRST_DATA_COL, src_col_end=LAST_DATA_COL,
                      min_row=2, max_row=None):
    if max_row is None:
        max_row = ws.max_row
    offset = dst_col_start - src_col_start

    for r in range(min_row, max_row + 1):
        for c in range(src_col_start, src_col_end + 1):
            src_cell = ws.cell(row=r, column=c)
            dst_cell = ws.cell(row=r, column=c + offset)
            dst_cell.value = src_cell.value
            copy_cell_style(src_cell, dst_cell)

    for merge in list(ws.merged_cells.ranges):
        if merge.min_col >= src_col_start and merge.max_col <= src_col_end and merge.min_row >= min_row:
            ws.merge_cells(start_row=merge.min_row, start_column=merge.min_col + offset,
                            end_row=merge.max_row, end_column=merge.max_col + offset)

    for c in range(src_col_start, src_col_end + 1):
        src_letter = get_column_letter(c)
        dst_letter = get_column_letter(c + offset)
        if src_letter in ws.column_dimensions:
            ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width


def set_week_dates(ws, col_start, monday_date, date_row=3):
    for i in range(7):
        cell = ws.cell(row=date_row, column=col_start + i)
        cell.value = datetime(monday_date.year, monday_date.month, monday_date.day) + timedelta(days=i)


def format_title(monday_date, sunday_date, prefix="RIVERWAY ROOM SCHEDULE"):
    if monday_date.month == sunday_date.month:
        return f"{prefix} - {monday_date.strftime('%B %d')} - {sunday_date.strftime('%d, %Y')}"
    elif monday_date.year == sunday_date.year:
        return f"{prefix} - {monday_date.strftime('%B %d')} - {sunday_date.strftime('%B %d, %Y')}"
    else:
        return f"{prefix} - {monday_date.strftime('%B %d, %Y')} - {sunday_date.strftime('%B %d, %Y')}"


def resolve_writable_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if type(cell).__name__ != 'MergedCell':
        return cell
    for merge in ws.merged_cells.ranges:
        if merge.min_row <= row <= merge.max_row and merge.min_col <= col <= merge.max_col:
            return ws.cell(row=merge.min_row, column=merge.min_col)
    return cell


def expand_occurrences(course):
    days_codes = [ch for ch in course['days'] if not ch.isspace()]
    target_weekdays = {weekday_for_code(ch) for ch in days_codes}
    d = course['start']
    while d <= course['end']:
        if d.weekday() in target_weekdays:
            yield d
        d += timedelta(days=1)


def build_cell_text(course, start_time, end_time):
    return f"{course['name']} - #{course['course_num']} - {course['instructor']} - ({start_time}-{end_time})"


def enable_wrap_text_everywhere(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                existing = cell.alignment
                cell.alignment = Alignment(
                    horizontal=existing.horizontal,
                    vertical=existing.vertical,
                    wrap_text=True,
                    text_rotation=existing.text_rotation,
                    indent=existing.indent,
                )


# ----------------------------------------------------------------------------
# MAIN FILL ROUTINE - driven by a pre-built "assignments" list, no input()
# ----------------------------------------------------------------------------
def build_week_structure(wb, template_ws, weeks, mode):
    """
    mode: 'sheets' or 'columns'
    Returns (week_sheet_map, week_col_map, main_ws)
      - Sheet mode populates week_sheet_map, leaves week_col_map empty, main_ws=None
      - Column mode populates week_col_map + main_ws, leaves week_sheet_map empty
    """
    week_sheet_map = {}
    week_col_map = {}
    main_ws = None

    if mode == 'sheets':
        for i, wk_monday in enumerate(weeks):
            sunday = wk_monday + timedelta(days=6)
            if i == 0 and 'main' in wb.sheetnames:
                ws = wb['main']
            else:
                ws = wb.copy_worksheet(template_ws)
                sheet_name = (wk_monday.strftime("%b%d") + "-" + sunday.strftime("%b%d")).replace(" ", "")[:31]
                base_name, n = sheet_name, 2
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{base_name}_{n}"[:31]
                    n += 1
                ws.title = sheet_name
            set_week_dates(ws, col_start=FIRST_DATA_COL, monday_date=wk_monday)
            ws['A1'] = format_title(wk_monday, sunday)
            week_sheet_map[wk_monday] = ws
    else:
        ws = wb['main'] if 'main' in wb.sheetnames else wb.copy_worksheet(template_ws)
        for i, wk_monday in enumerate(weeks):
            col_start = FIRST_DATA_COL + i * BLOCK_WIDTH
            if i > 0:
                clone_week_block(ws, dst_col_start=col_start)
            set_week_dates(ws, col_start=col_start, monday_date=wk_monday)
            week_col_map[wk_monday] = col_start
        last_col = FIRST_DATA_COL + len(weeks) * BLOCK_WIDTH - 1
        ws.unmerge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        ws['A1'] = format_title(weeks[0], weeks[-1] + timedelta(days=6))
        main_ws = ws

    return week_sheet_map, week_col_map, main_ws


def fill_schedule(mode, data_rows, assignments, rooms_by_name, week_sheet_map, week_col_map, main_ws):
    """
    assignments: dict keyed by course['row'] -> {
        'room_name': str, 'day_or_evening': 'd'/'e', 'start_time': str, 'end_time': str
    }
    rooms_by_name: dict room_name -> room_dict (from parse_room_schema)
    Returns list of conflicts: (sheet_title, cell_coord, existing_value, course_name)
    """
    conflicts = []
    placed_counts = {}

    for course in data_rows:
        assignment = assignments.get(course['row'])
        if not assignment:
            continue
        room = rooms_by_name.get(assignment['room_name'])
        if room is None:
            continue
        row_idx = room['day_row'] if assignment['day_or_evening'] == 'd' else room['evening_row']
        if row_idx is None:
            continue

        placed = 0
        for occ_date in expand_occurrences(course):
            wk_monday = monday_of(occ_date)
            if mode == 'sheets':
                target_ws = week_sheet_map.get(wk_monday)
                col_start = FIRST_DATA_COL
            else:
                target_ws = main_ws
                col_start = week_col_map.get(wk_monday)
            if target_ws is None or col_start is None:
                continue
            col = col_start + occ_date.weekday()
            cell = resolve_writable_cell(target_ws, row_idx, col)
            if cell.value not in (None, ""):
                conflicts.append((target_ws.title, cell.coordinate, str(cell.value), course['name']))
                continue
            cell.value = build_cell_text(course, assignment['start_time'], assignment['end_time'])
            placed += 1
        placed_counts[course['row']] = placed

    return conflicts, placed_counts
