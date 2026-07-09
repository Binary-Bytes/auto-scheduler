import io

import streamlit as st
import pandas as pd
import openpyxl

from backend import (
    parse_room_schema, load_data_rows, detect_base_year, build_week_list,
    build_week_structure, fill_schedule, suggest_room, enable_wrap_text_everywhere,
)

st.set_page_config(page_title="Riverway Room Schedule Auto-Filler", layout="wide")
st.title("📅 Riverway Room Schedule Auto-Filler")
st.caption("Upload the term's course list, review the room/time assignments, and download a filled-in schedule.")

# ----------------------------------------------------------------------------
# Step 1: Upload
# ----------------------------------------------------------------------------
uploaded_file = st.file_uploader(
    "Upload the workbook (must contain a 'Data' sheet and a 'Template' or 'main' sheet)",
    type=["xlsx"],
)

if uploaded_file is None:
    st.info("Waiting for a workbook to be uploaded...")
    st.stop()

# Parse once per uploaded file, cache in session_state
if "wb_bytes" not in st.session_state or st.session_state.get("uploaded_name") != uploaded_file.name:
    st.session_state["wb_bytes"] = uploaded_file.getvalue()
    st.session_state["uploaded_name"] = uploaded_file.name
    st.session_state.pop("assignments_df", None)  # reset any prior edits for a new file

try:
    wb = openpyxl.load_workbook(io.BytesIO(st.session_state["wb_bytes"]))
except Exception as e:
    st.error(f"Could not open this file as an Excel workbook: {e}")
    st.stop()

if "Data" not in wb.sheetnames:
    st.error("No sheet named 'Data' found. Please make sure your course list sheet is named exactly 'Data'.")
    st.stop()

TEMPLATE_SHEET_NAME = "Template" if "Template" in wb.sheetnames else "main"
if TEMPLATE_SHEET_NAME not in wb.sheetnames:
    st.error("No 'Template' or 'main' sheet found - one of these is needed as the room-grid layout to copy.")
    st.stop()

template_ws = wb[TEMPLATE_SHEET_NAME]

try:
    rooms = parse_room_schema(template_ws)
except Exception as e:
    st.error(f"Could not read the room layout from '{TEMPLATE_SHEET_NAME}': {e}")
    st.stop()

room_names = [r["room_name"] for r in rooms]
rooms_by_name = {r["room_name"]: r for r in rooms}
base_year = detect_base_year(template_ws)

try:
    data_rows = load_data_rows(wb["Data"], base_year)
except Exception as e:
    st.error(f"Could not read the 'Data' sheet: {e}")
    st.stop()

if not data_rows:
    st.warning("No course rows found in the 'Data' sheet.")
    st.stop()

min_start = min(d["start"] for d in data_rows)
max_end = max(d["end"] for d in data_rows)
weeks = build_week_list(min_start, max_end)

st.success(
    f"Loaded **{len(data_rows)}** course rows from '{TEMPLATE_SHEET_NAME}' baseline year **{base_year}**. "
    f"Schedule spans **{min_start} → {max_end}** ({len(weeks)} weeks)."
)

with st.expander("Rooms detected in the template"):
    for r in rooms:
        slots = []
        if r["day_row"]:
            slots.append("Day")
        if r["evening_row"]:
            slots.append("Evening")
        hint = f"  (hints: {', '.join(r['labels'][1:])})" if len(r["labels"]) > 1 else ""
        st.write(f"- **{r['room_name']}** [{'/'.join(slots)}]{hint}")

# ----------------------------------------------------------------------------
# Step 2: Layout mode
# ----------------------------------------------------------------------------
st.subheader("1. Choose the layout")
mode_label = st.radio(
    "How should the extra weeks be laid out?",
    ["Sheet Mode - one new tab per week", "Column Mode - one wide 'main' sheet"],
    horizontal=False,
)
mode = "sheets" if mode_label.startswith("Sheet") else "columns"

# ----------------------------------------------------------------------------
# Step 3: Editable assignment table
# ----------------------------------------------------------------------------
st.subheader("2. Review & fill in room / time for each course")
st.caption(
    "Room is auto-suggested where possible by matching the course name against the template's hint labels. "
    "Rows with no suggestion are left blank - please pick a room for those before generating."
)

if "assignments_df" not in st.session_state:
    records = []
    for course in data_rows:
        room, matched_label = suggest_room(course["name"], rooms)
        records.append({
            "Course Row": course["row"],
            "Course #": course["course_num"],
            "Name": course["name"],
            "Instructor": course["instructor"],
            "Start Date": str(course["start"]),
            "End Date": str(course["end"]),
            "Days": course["days"],
            "Notes": course["notes"],
            "Room": room["room_name"] if room else None,
            "Day/Evening": None,
            "Start Time": "",
            "End Time": "",
        })
    st.session_state["assignments_df"] = pd.DataFrame(records)

edited_df = st.data_editor(
    st.session_state["assignments_df"],
    column_config={
        "Course Row": st.column_config.NumberColumn(disabled=True),
        "Course #": st.column_config.TextColumn(disabled=True),
        "Name": st.column_config.TextColumn(disabled=True),
        "Instructor": st.column_config.TextColumn(disabled=True),
        "Start Date": st.column_config.TextColumn(disabled=True),
        "End Date": st.column_config.TextColumn(disabled=True),
        "Days": st.column_config.TextColumn(disabled=True),
        "Notes": st.column_config.TextColumn(disabled=True),
        "Room": st.column_config.SelectboxColumn(options=room_names, required=True),
        "Day/Evening": st.column_config.SelectboxColumn(options=["Day", "Evening"], required=True),
        "Start Time": st.column_config.TextColumn(help="e.g. 8:30am"),
        "End Time": st.column_config.TextColumn(help="e.g. 4:30pm"),
    },
    disabled=False,
    hide_index=True,
    use_container_width=True,
    key="assignment_editor",
)
st.session_state["assignments_df"] = edited_df

# ----------------------------------------------------------------------------
# Step 4: Generate
# ----------------------------------------------------------------------------
st.subheader("3. Generate the filled schedule")

missing = edited_df[
    edited_df["Room"].isna() | edited_df["Day/Evening"].isna()
    | (edited_df["Start Time"].str.strip() == "") | (edited_df["End Time"].str.strip() == "")
]

if len(missing) > 0:
    st.warning(f"{len(missing)} course row(s) are still missing Room / Day-Evening / times. Fill those in above before generating.")

generate_clicked = st.button("Generate Filled Schedule", type="primary", disabled=len(missing) > 0)

if generate_clicked:
    with st.spinner("Building the schedule..."):
        wb_out = openpyxl.load_workbook(io.BytesIO(st.session_state["wb_bytes"]))
        template_ws_out = wb_out[TEMPLATE_SHEET_NAME]
        data_rows_out = load_data_rows(wb_out["Data"], base_year)

        week_sheet_map, week_col_map, main_ws = build_week_structure(wb_out, template_ws_out, weeks, mode)

        assignments = {}
        for _, row in edited_df.iterrows():
            assignments[int(row["Course Row"])] = {
                "room_name": row["Room"],
                "day_or_evening": "d" if row["Day/Evening"] == "Day" else "e",
                "start_time": row["Start Time"],
                "end_time": row["End Time"],
            }

        rooms_by_name_out = {r["room_name"]: r for r in parse_room_schema(template_ws_out)}
        conflicts, placed_counts = fill_schedule(
            mode, data_rows_out, assignments, rooms_by_name_out, week_sheet_map, week_col_map, main_ws
        )
        enable_wrap_text_everywhere(wb_out)

        out_buffer = io.BytesIO()
        wb_out.save(out_buffer)
        out_buffer.seek(0)

    total_placed = sum(placed_counts.values())
    st.success(f"Done! Placed {total_placed} class occurrence(s) across {len(data_rows_out)} course(s).")

    if conflicts:
        st.error(f"{len(conflicts)} conflict(s) found - these cells already had content, so nothing was overwritten:")
        conflict_df = pd.DataFrame(conflicts, columns=["Sheet", "Cell", "Existing Content", "Course That Couldn't Be Placed"])
        st.dataframe(conflict_df, use_container_width=True)
    else:
        st.info("No conflicts - every course occurrence was placed successfully.")

    st.download_button(
        "⬇️ Download Filled Schedule (.xlsx)",
        data=out_buffer,
        file_name="Room_Schedule_Filled.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
