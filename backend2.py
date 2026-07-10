"""
Pure backend logic for the Night School Weekday Schedule Auto-Filler.

Unlike the room-schedule tool, there is no pre-built template grid to match
against here: each weekday ("Monday", "Tuesday", ...) gets its own sheet,
and every course gets its own ROW within the sheet(s) for the weekday(s) it
meets on. Weeks are single columns (one column per calendar occurrence of
that weekday), grouped under a month header, and a course's active weeks
are shaded light gray rather than filled with text.
"""
import re
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

WEEKDAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
CODE_TO_WEEKDAY = {'M': 0, 'T': 1, 'W': 2, 'R': 3, 'F': 4, 'S': 5}  # anything else -> 6 (Sunday)

HEADER_COLS = ["Course", "Course #", "Room", "Instructor", "Start", "End", "Time", "NOTE"]
FIRST_WEEK_COL = 9   # column I - first week-of-term column, matching the real file's layout
TITLE_ROW = 1
HEADER_ROW = 2
DATE_ROW = 3
FIRST_COURSE_ROW = 4

# ----------------------------------------------------------------------------
# STYLE CONSTANTS
# ----------------------------------------------------------------------------
TITLE_FONT = Font(name="Calibri", bold=True, size=14)
HEADER_FONT = Font(name="Calibri", bold=True, size=12)
MONTH_FONT = Font(name="Calibri", bold=True, size=11)
DATE_FONT = Font(name="Calibri", bold=True, size=10)
BODY_FONT = Font(name="Calibri", size=11)
LEGEND_FONT = Font(name="Calibri", italic=True, size=9)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_SIDE = Side(style="thin", color="BFBFBF")
THICK_SIDE = Side(style="medium", color="000000")  # Clean, pronounced month separator
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

HEADER_FILL = PatternFill(patternType="solid", fgColor="D9D9D9")
MONTH_FILL = PatternFill(patternType="solid", fgColor="EDEDED")

WEEK_COL_WIDTH = 4.3   # narrow, near-square columns for the week-date grid
DATE_ROW_HEIGHT = 16
HEADER_ROW_HEIGHT = 20
BODY_ROW_HEIGHT = 28


def make_shade_fill():
    # Matches the "Background 1, Darker 15%" preset used in the real reference file
    return PatternFill(patternType="solid", fgColor=Color(theme=0, tint=-0.1499984740745262, type="theme"))


def weekday_for_code(ch):
    return CODE_TO_WEEKDAY.get(ch.upper(), 6)


def weekdays_for_days_string(days_str):
    """Returns the set of weekday indices (0=Mon..6=Sun) a Days code string covers."""
    return {weekday_for_code(ch) for ch in days_str if not ch.isspace()}


# ----------------------------------------------------------------------------
# DATA SHEET LOADING (same conventions as the room-schedule tool)
# ----------------------------------------------------------------------------
def parse_data_date(value, base_year, fall_start_month=7):
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'year') and hasattr(value, 'month'):
        return value
    s = str(value).strip()
    m = re.match(r'^(\d{1,2})[-\s]([A-Za-z]{3,})$', s)
    if not m:
        raise ValueError(f"Could not parse date value: {value!r}")
    day = int(m.group(1))
    mon_str = m.group(2)[:3].title()
    month = datetime.strptime(mon_str, "%b").month
    year = base_year if month >= fall_start_month else base_year + 1
    return datetime(year, month, day).date()


def load_data_rows(ws_data):
    headers = {}
    for c in range(1, ws_data.max_column + 1):
        v = ws_data.cell(row=1, column=c).value
        if v:
            headers[str(v).strip().lower()] = c

    required = ['starts', 'ends', 'course', 'name', 'days', 'instructor']
    for req in required:
        if req not in headers:
            raise ValueError(f"Data sheet is missing required column: '{req}'")

    first_start_raw = None
    for r in range(2, ws_data.max_row + 1):
        v = ws_data.cell(row=r, column=headers['starts']).value
        if v is not None:
            first_start_raw = v
            break
    if first_start_raw is None:
        raise ValueError("Could not find any Start dates in the Data sheet.")
    if hasattr(first_start_raw, 'year'):
        base_year = first_start_raw.year
    else:
        base_year = datetime.now().year

    rows = []
    for r in range(2, ws_data.max_row + 1):
        name_val = ws_data.cell(row=r, column=headers['name']).value
        if name_val is None or str(name_val).strip() == "":
            continue

        start_raw = ws_data.cell(row=r, column=headers['starts']).value
        end_raw = ws_data.cell(row=r, column=headers['ends']).value
        start_date = parse_data_date(start_raw, base_year)
        end_date = parse_data_date(end_raw, base_year)

        course_raw = str(ws_data.cell(row=r, column=headers['course']).value or "").strip()
        days_raw = str(ws_data.cell(row=r, column=headers['days']).value or "").strip()
        instructor = str(ws_data.cell(row=r, column=headers['instructor']).value or "").strip()
        name = str(name_val).strip()

        notes = ""
        if 'notes' in headers:
            notes = str(ws_data.cell(row=r, column=headers['notes']).value or "").strip()

        rows.append({
            'row': r,
            'start': start_date,
            'end': end_date,
            'course_raw': course_raw,
            'name': name,
            'days': days_raw,
            'instructor': instructor,
            'notes': notes,
        })
    return rows


def required_weekdays(data_rows):
    """Which of the 7 weekdays have at least one course meeting on them."""
    needed = set()
    for course in data_rows:
        needed |= weekdays_for_days_string(course['days'])
    return sorted(needed)  # 0=Mon..6=Sun


# ----------------------------------------------------------------------------
# WEEK-COLUMN CALCULATION (shared term range across all weekday sheets)
# ----------------------------------------------------------------------------
def term_date_range(data_rows):
    """
    Returns the expanded date range stretching from the 1st day of the earliest 
    month to the final day of the latest month across all courses.
    """
    if not data_rows:
        return None, None
    min_start = min(c['start'] for c in data_rows)
    max_end = max(c['end'] for c in data_rows)
    
    # Expand to the 1st day of the earliest start month
    start_bound = min_start.replace(day=1)
    
    # Expand to the last day of the latest end month
    if max_end.month == 12:
        end_bound = max_end.replace(year=max_end.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        end_bound = max_end.replace(month=max_end.month + 1, day=1) - timedelta(days=1)
        
    return start_bound, end_bound


def week_dates_for_weekday(weekday_idx, term_min_start, term_max_end):
    if term_min_start is None or term_max_end is None:
        return []

    # first occurrence of this weekday on/after term_min_start boundary
    delta = (weekday_idx - term_min_start.weekday()) % 7
    d = term_min_start + timedelta(days=delta)
    dates = []
    while d <= term_max_end:
        dates.append(d)
        d += timedelta(days=7)
    return dates


def group_dates_by_month(dates):
    """Returns list of (month_label, [dates]) preserving order, for the header row."""
    groups = []
    for d in dates:
        label = d.strftime("%B").upper()
        if groups and groups[-1][0] == label:
            groups[-1][1].append(d)
        else:
            groups.append((label, [d]))
    return groups


# ----------------------------------------------------------------------------
# WORKBOOK CONSTRUCTION
# ----------------------------------------------------------------------------
def build_sheet_for_weekday(wb, weekday_idx, dates, school_name="Night School"):
    weekday_name = WEEKDAY_NAMES[weekday_idx]
    ws = wb.create_sheet(title=weekday_name)

    year_label = dates[0].year if dates else datetime.now().year
    last_col = FIRST_WEEK_COL + max(len(dates) - 1, 0)

    title_cell = ws.cell(row=TITLE_ROW, column=1,
                          value=f"{school_name} \u2013 Room Schedule \u2013 FALL {year_label} \u2013 {weekday_name}")
    ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=max(last_col, 8))
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER
    ws.row_dimensions[TITLE_ROW].height = 24

    for i, label in enumerate(HEADER_COLS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=i, value=label)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # month header + date row
    col = FIRST_WEEK_COL
    for month_label, month_dates in group_dates_by_month(dates):
        start_col = col
        for d in month_dates:
            dcell = ws.cell(row=DATE_ROW, column=col, value=d.day)
            dcell.font = DATE_FONT
            dcell.alignment = CENTER
            dcell.border = THIN_BORDER
            hcell = ws.cell(row=HEADER_ROW, column=col)
            hcell.fill = MONTH_FILL
            hcell.border = THIN_BORDER
            col += 1
        end_col = col - 1
        cell = ws.cell(row=HEADER_ROW, column=start_col, value=month_label)
        cell.font = MONTH_FONT
        cell.alignment = CENTER
        cell.fill = MONTH_FILL
        cell.border = THIN_BORDER
        if end_col > start_col:
            ws.merge_cells(start_row=HEADER_ROW, start_column=start_col, end_row=HEADER_ROW, end_column=end_col)

    ws.row_dimensions[HEADER_ROW].height = HEADER_ROW_HEIGHT
    ws.row_dimensions[DATE_ROW].height = DATE_ROW_HEIGHT

    return ws


def add_course_row(ws, row_idx, course, room, time_str, dates_for_weekday, weekday_idx):
    days_suffix = f" ({course['days']})" if course['days'] else ""

    values = [
        course['name'] + days_suffix,
        course['course_raw'],
        room,
        course['instructor'],
        datetime(course['start'].year, course['start'].month, course['start'].day),
        datetime(course['end'].year, course['end'].month, course['end'].day),
        time_str,
        course['notes'],
    ]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = LEFT_WRAP if col_idx in (1, 8) else CENTER

    ws.cell(row=row_idx, column=5).number_format = "d-mmm-yyyy"
    ws.cell(row=row_idx, column=6).number_format = "d-mmm-yyyy"

    ws.row_dimensions[row_idx].height = BODY_ROW_HEIGHT

    shade = make_shade_fill()
    placed = 0
    for i, d in enumerate(dates_for_weekday):
        col = FIRST_WEEK_COL + i
        wcell = ws.cell(row=row_idx, column=col)
        wcell.border = THIN_BORDER
        if course['start'] <= d <= course['end']:
            wcell.fill = shade
            placed += 1
    return placed


def add_legend(ws, row_idx):
    cell = ws.cell(row=row_idx, column=1, value="S = Start of course, X = End of course, E = School closed "
                                                 "(enter manually on the relevant week as needed)")
    cell.font = LEGEND_FONT


def enable_wrap_text_everywhere(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                existing = cell.alignment
                cell.alignment = Alignment(
                    horizontal=existing.horizontal,
                    vertical=existing.vertical or "center",
                    wrap_text=True,
                    text_rotation=existing.text_rotation,
                    indent=existing.indent,
                )


def style_columns(ws, last_week_col):
    fixed_widths = {1: 26, 2: 12, 3: 8, 4: 16, 5: 13, 6: 13, 7: 11, 8: 22}
    for col, width in fixed_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    for col in range(FIRST_WEEK_COL, last_week_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = WEEK_COL_WIDTH


def build_workbook(data_rows, assignments, school_name="Night School"):
    wb = Workbook()
    wb.remove(wb.active)

    needed_weekdays = required_weekdays(data_rows)
    placed_counts = {}

    term_min_start, term_max_end = term_date_range(data_rows)

    for weekday_idx in needed_weekdays:
        dates = week_dates_for_weekday(weekday_idx, term_min_start, term_max_end)
        ws = build_sheet_for_weekday(wb, weekday_idx, dates, school_name=school_name)

        row_idx = FIRST_COURSE_ROW
        for course in data_rows:
            if weekday_idx not in weekdays_for_days_string(course['days']):
                continue
            assignment = assignments.get(course['row'], {})
            room = assignment.get('room', '')
            time_str = assignment.get('time', '')
            placed = add_course_row(ws, row_idx, course, room, time_str, dates, weekday_idx)
            placed_counts[(course['row'], weekday_idx)] = placed
            row_idx += 1

        add_legend(ws, row_idx + 1)

        last_week_col = FIRST_WEEK_COL + max(len(dates) - 1, 0)
        style_columns(ws, last_week_col)
        
        # --------------------------------------------------------------------
        # Stylistic Pass: Apply Thick Right Border Between Distinct Months
        # --------------------------------------------------------------------
        curr_col = FIRST_WEEK_COL
        month_groups = group_dates_by_month(dates)
        
        # We don't need a right border on the very last month column group
        for _, month_dates in month_groups[:-1]:
            curr_col += len(month_dates)
            target_col_idx = curr_col - 1
            
            # Apply down through headers, dates, and all scheduled data rows
            for r in range(HEADER_ROW, row_idx):
                cell = ws.cell(row=r, column=target_col_idx)
                cell.border = Border(
                    left=cell.border.left,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                    right=THICK_SIDE
                )

        ws.sheet_view.showGridLines = False

    enable_wrap_text_everywhere(wb)
    return wb, placed_counts