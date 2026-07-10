import io

import streamlit as st
import pandas as pd
import openpyxl

from backend2 import load_data_rows, required_weekdays, build_workbook, WEEKDAY_NAMES

st.set_page_config(page_title="Night School Weekday Schedule Auto-Filler", layout="wide")
st.title("🌙 North Night School Scheduler")
st.caption(
    "Upload the course list, fill in Room + Time for each course, and get back one sheet "
    "per weekday with every class's active weeks shaded automatically."
)

# ----------------------------------------------------------------------------
# Step 1: Upload
# ----------------------------------------------------------------------------
uploaded_file = st.file_uploader(
    "Upload the workbook (must contain a 'Data' sheet with Starts/Ends/Course/Name/Days/Instructor columns)",
    type=["xlsx"],
)

if uploaded_file is None:
    st.info("Waiting for a workbook to be uploaded...")
    st.stop()

if "wb_bytes" not in st.session_state or st.session_state.get("uploaded_name") != uploaded_file.name:
    st.session_state["wb_bytes"] = uploaded_file.getvalue()
    st.session_state["uploaded_name"] = uploaded_file.name
    st.session_state.pop("assignments_df2", None)

try:
    wb_in = openpyxl.load_workbook(io.BytesIO(st.session_state["wb_bytes"]))
except Exception as e:
    st.error(f"Could not open this file as an Excel workbook: {e}")
    st.stop()

if "Data" not in wb_in.sheetnames:
    st.error("No sheet named 'Data' found. Please make sure your course list sheet is named exactly 'Data'.")
    st.stop()

try:
    data_rows = load_data_rows(wb_in["Data"])
except Exception as e:
    st.error(f"Could not read the 'Data' sheet: {e}")
    st.stop()

if not data_rows:
    st.warning("No course rows found in the 'Data' sheet.")
    st.stop()

needed_weekdays = required_weekdays(data_rows)
st.success(
    f"Loaded **{len(data_rows)}** course rows. "
    f"Sheets will be created for: **{', '.join(WEEKDAY_NAMES[i] for i in needed_weekdays)}**."
)

school_name = st.text_input("School name (used in each sheet's title bar)", value="Night School")

# ----------------------------------------------------------------------------
# Step 2: Editable Room / Time table (one row per course - not per weekday,
# since a course meeting on multiple days uses the same room & time on all of them)
# ----------------------------------------------------------------------------
st.subheader("1. Fill in Room and Time for each course")
st.caption("These apply to every weekday that course meets on (e.g. a Tue/Thu course only needs one Room + Time).")

if "assignments_df2" not in st.session_state:
    records = []
    for course in data_rows:
        records.append({
            "Course Row": course["row"],
            "Course #": course["course_raw"],
            "Name": course["name"],
            "Days": course["days"],
            "Instructor": course["instructor"],
            "Start Date": str(course["start"]),
            "End Date": str(course["end"]),
            "Notes": course["notes"],
            "Room": "",
            "Time": "",
        })
    st.session_state["assignments_df2"] = pd.DataFrame(records)

with st.form("assignment_form2"):
    edited_df = st.data_editor(
        st.session_state["assignments_df2"],
        column_config={
            "Course Row": st.column_config.NumberColumn(disabled=True),
            "Course #": st.column_config.TextColumn(disabled=True),
            "Name": st.column_config.TextColumn(disabled=True),
            "Days": st.column_config.TextColumn(disabled=True),
            "Instructor": st.column_config.TextColumn(disabled=True),
            "Start Date": st.column_config.TextColumn(disabled=True),
            "End Date": st.column_config.TextColumn(disabled=True),
            "Notes": st.column_config.TextColumn(disabled=True),
            "Room": st.column_config.TextColumn(help="e.g. 2104"),
            "Time": st.column_config.TextColumn(help="e.g. 6:15-9:15"),
        },
        hide_index=True,
        use_container_width=True,
        key="assignment_editor2",
    )
    generate_clicked = st.form_submit_button("Generate Weekday Schedules", type="primary")

st.session_state["assignments_df2"] = edited_df

# ----------------------------------------------------------------------------
# Step 3: Generate
# ----------------------------------------------------------------------------
st.subheader("2. Generate")

if generate_clicked:
    missing = edited_df[(edited_df["Room"].str.strip() == "") | (edited_df["Time"].str.strip() == "")]
    if len(missing) > 0:
        st.warning(f"{len(missing)} course row(s) are still missing Room / Time. Fill those in above and click Generate again.")
        st.stop()

    with st.spinner("Building the schedule..."):
        assignments = {}
        for _, row in edited_df.iterrows():
            assignments[int(row["Course Row"])] = {"room": row["Room"], "time": row["Time"]}

        wb_out, placed_counts = build_workbook(data_rows, assignments, school_name=school_name)

        out_buffer = io.BytesIO()
        wb_out.save(out_buffer)
        out_buffer.seek(0)

    total_placed = sum(placed_counts.values())
    st.success(
        f"Done! Created {len(wb_out.sheetnames)} weekday sheet(s) "
        f"({', '.join(wb_out.sheetnames)}) and shaded {total_placed} class-week cell(s) total."
    )
    st.info(
        "Remember: the S / X / E annotations (start / end / school-closed) are not filled in "
        "automatically - add those by hand on the relevant week once the file is downloaded."
    )

    st.download_button(
        "⬇️ Download Weekday Schedule (.xlsx)",
        data=out_buffer,
        file_name="Night_School_Weekday_Schedule.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
